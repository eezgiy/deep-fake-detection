import argparse
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader
from torchvision import datasets, transforms
from torchvision import models
from tqdm import tqdm
from sklearn.metrics import precision_score, recall_score, f1_score
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

# Argparse kullanarak komut satırından veri seti yolu alıyoruz
parser = argparse.ArgumentParser(description="Deep Fake Detection using CNN with PyTorch")
parser.add_argument('--dataset', type=str, required=True, help="Path to the dataset")
parser.add_argument('--epochs', type=int, default=20, help="Number of epochs")
args = parser.parse_args()

# Veri seti yolunu alıyoruz
train_dir = args.dataset + '/cnn_train'  # Eğitim verisi yolu
test_dir = args.dataset + '/cnn_test'    # Test verisi yolu

print(f"Training data directory: {train_dir}")
print(f"Test data directory: {test_dir}")

# Veri ön işleme ve augmentasyon
train_transform = transforms.Compose([
    transforms.Resize((64, 64)),
    transforms.RandomHorizontalFlip(),
    transforms.RandomRotation(20),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])  # ImageNet'in ortalama ve std'si
])

test_transform = transforms.Compose([
    transforms.Resize((64, 64)),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
])

# Veri setini yükleme
train_dataset = datasets.ImageFolder(train_dir, transform=train_transform)
test_dataset = datasets.ImageFolder(test_dir, transform=test_transform)

# DataLoader kullanarak veri yükleyicileri oluşturuyoruz
train_loader = DataLoader(train_dataset, batch_size=32, shuffle=True)
test_loader = DataLoader(test_dataset, batch_size=32, shuffle=False)

# CNN modelini oluşturuyoruz (ResNet18'i örnek alarak)
class DeepFakeDetector(nn.Module):
    def __init__(self):
        super(DeepFakeDetector, self).__init__()
        self.model = models.resnet18(pretrained=True)
        self.model.fc = nn.Linear(self.model.fc.in_features, 1)  # Çıktı boyutunu 1'e indiriyoruz

    def forward(self, x):
        return torch.sigmoid(self.model(x))

# Modeli oluşturuyoruz
model = DeepFakeDetector()

# Cihaz (GPU veya CPU) seçimi
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
model = model.to(device)

# Kayıp fonksiyonu ve optimizer
criterion = nn.BCELoss()  # Binary Cross-Entropy Loss
optimizer = optim.Adam(model.parameters(), lr=0.0001)

# Eğitim döngüsü
def train_model():
    model.train()
    running_loss = 0.0
    correct_preds = 0
    total_preds = 0
    all_preds = []
    all_labels = []

    for inputs, labels in tqdm(train_loader):
        inputs, labels = inputs.to(device), labels.to(device)

        optimizer.zero_grad()

        outputs = model(inputs)
        loss = criterion(outputs.view(-1), labels.float())  # BCELoss expects labels as float

        loss.backward()
        optimizer.step()

        running_loss += loss.item()

        # Doğru tahminleri hesaplıyoruz
        predicted = (outputs > 0.5).float()
        correct_preds += (predicted == labels).sum().item()
        total_preds += labels.size(0)

        # Tahminleri ve etiketleri topluyoruz
        all_preds.extend(predicted.cpu().numpy())
        all_labels.extend(labels.cpu().numpy())

    avg_loss = running_loss / len(train_loader)
    accuracy = correct_preds / total_preds

    # Precision, Recall, F1 Score
    precision = precision_score(all_labels, all_preds)
    recall = recall_score(all_labels, all_preds)
    f1 = f1_score(all_labels, all_preds)

    return avg_loss, accuracy, precision, recall, f1

# Test döngüsü
def test_model():
    model.eval()
    running_loss = 0.0
    correct_preds = 0
    total_preds = 0
    all_preds = []
    all_labels = []

    with torch.no_grad():
        for inputs, labels in tqdm(test_loader):
            inputs, labels = inputs.to(device), labels.to(device)

            outputs = model(inputs)
            loss = criterion(outputs.view(-1), labels.float())

            running_loss += loss.item()

            # Doğru tahminleri hesaplıyoruz
            predicted = (outputs > 0.5).float()
            correct_preds += (predicted == labels).sum().item()
            total_preds += labels.size(0)

            # Tahminleri ve etiketleri topluyoruz
            all_preds.extend(predicted.cpu().numpy())
            all_labels.extend(labels.cpu().numpy())

    avg_loss = running_loss / len(test_loader)
    accuracy = correct_preds / total_preds

    # Precision, Recall, F1 Score
    precision = precision_score(all_labels, all_preds)
    recall = recall_score(all_labels, all_preds)
    f1 = f1_score(all_labels, all_preds)

    return avg_loss, accuracy, precision, recall, f1

# Çıktıları dosyaya yazmak için
def save_metrics_to_excel(epoch, train_loss, train_accuracy, train_precision, train_recall, train_f1, 
                         test_loss, test_accuracy, test_precision, test_recall, test_f1, filename="metrics.xlsx"):
    # Metrikleri bir sözlükte topluyoruz
    metrics = {
        'Epoch': [epoch],
        'Train Loss': [round(train_loss, 2)],
        'Train Accuracy': [round(train_accuracy, 2)],
        'Train Precision': [round(train_precision, 2)],
        'Train Recall': [round(train_recall, 2)],
        'Train F1 Score': [round(train_f1, 2)],
        'Test Loss': [round(test_loss, 2)],
        'Test Accuracy': [round(test_accuracy, 2)],
        'Test Precision': [round(test_precision, 2)],
        'Test Recall': [round(test_recall, 2)],
        'Test F1 Score': [round(test_f1, 2)]
    }

    # Pandas DataFrame oluşturuyoruz
    df = pd.DataFrame(metrics)

    # Excel dosyasını kontrol ediyoruz
    try:
        # Dosyayı açıyoruz, eğer varsa
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = book
        df.to_excel(writer, index=False, header=False, startrow=book.active.max_row)
        writer.save()
    except FileNotFoundError:
        # Dosya yoksa yeni oluşturuyoruz
        df.to_excel(filename, index=False)

# Modeli eğitiyoruz ve doğruluyoruz
num_epochs = 20
for epoch in range(num_epochs):
    print(f"Epoch {epoch+1}/{num_epochs}")

    # Eğitim adımı
    train_loss, train_accuracy, train_precision, train_recall, train_f1 = train_model()

    # Test adımı
    test_loss, test_accuracy, test_precision, test_recall, test_f1 = test_model()

    # Çıktıları dosyaya kaydediyoruz
    save_metrics_to_excel(epoch, train_loss, train_accuracy, train_precision, train_recall, train_f1, 
                         test_loss, test_accuracy, test_precision, test_recall, test_f1)
    print(f"Epoch {epoch+1} metrikleri dosyaya kaydedildi.")
